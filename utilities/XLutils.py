import openpyxl


def ReadData(file, sheetname, rownum, colnumnum):
    workbook = openpyxl.load_workbook(file)  # file
    sheet = workbook[sheetname]  # sheet
    return sheet.cell(row=rownum, column=colnumnum).value


def WriteData(file, sheetname, rownum, colnumnum, data):
    workbook = openpyxl.load_workbook(file)  # file
    sheet = workbook[sheetname]  # sheet
    sheet.cell(row=rownum, column=colnumnum).value = data  # enter data
    workbook.save(file)  # Save the file


def RowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)  # file
    sheet = workbook[sheetname]  # sheet
    return sheet.max_row
